#!/usr/bin/env python3
"""Automate ChatGPT web in real Chrome, one workbook prompt at a time.

The browser-specific parts are intentionally isolated in this file so Codex can
port them to Gemini while preserving the workbook merge/progress/validation
logic.
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import subprocess
import sys
import time
import urllib.request
from dataclasses import asdict, dataclass
from pathlib import Path

try:
    from playwright.sync_api import Error as PlaywrightError
    from playwright.sync_api import Page, TimeoutError as PlaywrightTimeoutError, sync_playwright
except ModuleNotFoundError:
    print("Install dependencies first: python3 -m pip install -r requirements.txt", file=sys.stderr)
    raise SystemExit(1)

ROOT = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = ROOT / "ceo_fitness_tracker_v2_sequential.xlsx"
DEFAULT_PROGRESS = ROOT / "progress.json"
DEFAULT_OUTPUTS = ROOT / "chatgpt_outputs"
IMPORTANT_SHEET = "Important Data"
SOURCES_SHEET = "Sources Evidence"

FORMULA_HEADERS = {
    "Height Total Inches",
    "Calculated BMI",
    "BMI Category",
    "Formula-Estimated Body Fat %",
    "Formula BF% Category",
    "Apparent BMI Category",
    "Visual BF% Category",
    "Category QA Flag",
}
IMPORTANT_EDITABLE_HEADERS = {
    "Height Feet",
    "Height Inches",
    "Weight lbs",
    "Apparent / Visual BMI Final",
    "Visual Body Fat % Final",
    "Visual Leanness / Adiposity Score Final",
    "AI Model Used",
    "Status",
}

COMPOSER_SELECTORS = ["#prompt-textarea", 'textarea[data-testid="prompt-textarea"]', 'div[contenteditable="true"]', "textarea"]
SEND_SELECTORS = ['form button[data-testid="send-button"]', 'button[data-testid="send-button"]', 'button[aria-label*="Send" i]', 'button[data-testid*="send" i]']
STOP_SELECTORS = ['button[data-testid="stop-button"]', 'button[aria-label*="Stop" i]', 'button[data-testid*="stop" i]']
CONTEXT_LIMIT_PHRASES = ["context limit", "long conversation", "start a new chat", "conversation is too long", "maximum length"]
BREAK_PHRASES = ["take a break", "come back later", "you've been chatting", "you’ve been chatting"]
BLOCKING_PHRASES = ["too many requests", "rate limit", "verify you are human", "unusual activity", "are you human"]


@dataclass
class Progress:
    next_prompt: int
    latest_excel_file: str
    current_chat_start_prompt: int
    completed_prompts: list[int]
    updated_at: float


class BlockingNoticeError(RuntimeError):
    pass


def save_progress(path: Path, progress: Progress) -> None:
    path.write_text(json.dumps(asdict(progress), indent=2, sort_keys=True) + "\n")


def load_or_create_progress(args: argparse.Namespace) -> Progress:
    path = Path(args.progress)
    if path.exists() and args.resume:
        data = json.loads(path.read_text())
        return Progress(
            next_prompt=int(data["next_prompt"]),
            latest_excel_file=str(data["latest_excel_file"]),
            current_chat_start_prompt=int(data.get("current_chat_start_prompt", data["next_prompt"])),
            completed_prompts=[int(x) for x in data.get("completed_prompts", [])],
            updated_at=float(data.get("updated_at", time.time())),
        )
    initial = Path(args.initial_excel).expanduser().resolve()
    if not initial.exists():
        raise FileNotFoundError(f"Initial Excel file not found: {initial}")
    return Progress(args.start_prompt, str(initial), args.start_prompt, [], time.time())


def header_map(ws) -> dict[str, int]:
    return {str(ws.cell(1, c).value).strip(): c for c in range(1, (ws.max_column or 0) + 1) if ws.cell(1, c).value is not None}


def prompt_row_map(ws) -> dict[int, int]:
    headers = header_map(ws)
    prompt_col = headers.get("Prompt ID")
    rows: dict[int, int] = {}
    if not prompt_col:
        return rows
    for row in range(2, (ws.max_row or 0) + 1):
        value = ws.cell(row, prompt_col).value
        try:
            if value is not None:
                rows[int(value)] = row
        except (TypeError, ValueError):
            pass
    return rows


def copy_by_header(src_ws, dst_ws, prompt_id: int, allowed_headers: set[str] | None = None) -> int:
    src_headers = header_map(src_ws)
    dst_headers = header_map(dst_ws)
    src_rows = prompt_row_map(src_ws)
    dst_rows = prompt_row_map(dst_ws)
    if prompt_id not in src_rows or prompt_id not in dst_rows:
        return 0
    copied = 0
    for header, src_col in src_headers.items():
        if header in {"index", "Prompt ID"} or header not in dst_headers or header in FORMULA_HEADERS:
            continue
        if allowed_headers is not None and header not in allowed_headers:
            continue
        value = src_ws.cell(src_rows[prompt_id], src_col).value
        if value is not None:
            dst_ws.cell(dst_rows[prompt_id], dst_headers[header]).value = value
            copied += 1
    return copied


def validate_workbook_prompt_filled(wb, prompt_number: int, label: str) -> None:
    if IMPORTANT_SHEET not in wb.sheetnames or SOURCES_SHEET not in wb.sheetnames:
        raise RuntimeError(f"{label} does not contain required sheets.")
    imp = wb[IMPORTANT_SHEET]
    src = wb[SOURCES_SHEET]
    imp_rows = prompt_row_map(imp)
    src_rows = prompt_row_map(src)
    if prompt_number not in imp_rows or prompt_number not in src_rows:
        raise RuntimeError(f"{label} does not contain Prompt {prompt_number}.")
    ih = header_map(imp)
    sh = header_map(src)
    status = imp.cell(imp_rows[prompt_number], ih.get("Status", 24)).value
    source_url = src.cell(src_rows[prompt_number], sh.get("Height Source URL", 6)).value
    audit_q1 = src.cell(src_rows[prompt_number], sh.get("Audit Q1 | Filled only assigned row? (Y/N)", 62)).value
    if status in (None, "", "Not Started") or source_url in (None, "") or audit_q1 in (None, ""):
        raise RuntimeError(f"{label} did not fill Prompt {prompt_number}. Found Status={status!r}, Height Source URL={source_url!r}, Audit Q1={audit_q1!r}.")


def merge_download_into_canonical(downloaded_file: Path, canonical_file: Path, prompt_number: int, backup_dir: Path) -> Path:
    from openpyxl import load_workbook

    downloaded_file = downloaded_file.resolve()
    canonical_file = canonical_file.resolve()
    src_wb = load_workbook(downloaded_file, data_only=False)
    try:
        validate_workbook_prompt_filled(src_wb, prompt_number, f"Downloaded workbook {downloaded_file.name}")
        backup_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(canonical_file, backup_dir / f"before_prompt_{prompt_number:03d}_{int(time.time())}_{canonical_file.name}")
        dst_wb = load_workbook(canonical_file, data_only=False)
        try:
            important = copy_by_header(src_wb[IMPORTANT_SHEET], dst_wb[IMPORTANT_SHEET], prompt_number, IMPORTANT_EDITABLE_HEADERS)
            sources = copy_by_header(src_wb[SOURCES_SHEET], dst_wb[SOURCES_SHEET], prompt_number, None)
            if important == 0 and sources == 0:
                raise RuntimeError(f"No cells were merged for Prompt {prompt_number} from {downloaded_file}")
            dst_wb.save(canonical_file)
        finally:
            dst_wb.close()
    finally:
        src_wb.close()
    return canonical_file


def filled_prompt_numbers(canonical_file: Path) -> set[int]:
    from openpyxl import load_workbook

    wb = load_workbook(canonical_file, data_only=False, read_only=False)
    filled: set[int] = set()
    try:
        if IMPORTANT_SHEET not in wb.sheetnames or SOURCES_SHEET not in wb.sheetnames:
            return filled
        imp = wb[IMPORTANT_SHEET]
        src = wb[SOURCES_SHEET]
        ih = header_map(imp)
        sh = header_map(src)
        src_rows = prompt_row_map(src)
        for row in range(2, imp.max_row + 1):
            pid = imp.cell(row, ih.get("Prompt ID", 1)).value
            if not isinstance(pid, int) or pid not in src_rows:
                continue
            status = imp.cell(row, ih.get("Status", 24)).value
            source_url = src.cell(src_rows[pid], sh.get("Height Source URL", 6)).value
            audit_q1 = src.cell(src_rows[pid], sh.get("Audit Q1 | Filled only assigned row? (Y/N)", 62)).value
            if status not in (None, "", "Not Started") and source_url not in (None, "") and audit_q1 not in (None, ""):
                filled.add(pid)
    finally:
        wb.close()
    return filled


def page_text(page: Page) -> str:
    try:
        return page.locator("body").inner_text(timeout=3000).lower()
    except PlaywrightError:
        return ""


def phrase_visible(page: Page, phrases: list[str]) -> str | None:
    text = page_text(page)
    return next((phrase for phrase in phrases if phrase in text), None)


def is_conversation_page(page: Page) -> bool:
    return "chatgpt.com/c/" in (page.url or "")


def chatgpt_page(context) -> Page:
    for page in reversed(context.pages):
        if "chatgpt.com" in (page.url or ""):
            page.bring_to_front()
            return page
    page = context.new_page()
    page.goto("https://chatgpt.com/", wait_until="commit", timeout=15000)
    return page


def close_non_chatgpt_tabs(context, keep_page: Page | None = None) -> None:
    for page in list(context.pages):
        if page == keep_page:
            continue
        url = page.url or ""
        if "chatgpt.com" not in url and url != "about:blank":
            try:
                print(f"Closing non-ChatGPT tab: {url}")
                page.close()
            except PlaywrightError:
                pass


def install_tab_guard(context) -> None:
    def guard(page: Page) -> None:
        try:
            page.wait_for_load_state("domcontentloaded", timeout=3000)
            url = page.url or ""
            if "chatgpt.com" not in url and url != "about:blank":
                print(f"Closing external tab opened by automation: {url}")
                page.close()
        except PlaywrightError:
            pass
    context.on("page", guard)


def dismiss_transient_modal(page: Page) -> None:
    if phrase_visible(page, BLOCKING_PHRASES) and not phrase_visible(page, BREAK_PHRASES):
        return
    modal = page.locator("#modal-walnut-content, [data-testid='modal-walnut-content']").last
    try:
        modal_visible = modal.count() and modal.is_visible(timeout=500)
    except PlaywrightError:
        modal_visible = False
    if phrase_visible(page, BREAK_PHRASES) or modal_visible:
        for locator in [
            page.get_by_role("button", name=re.compile(r"ok|got it|continue|keep going|dismiss|close|resume|try again|not now|later", re.I)),
            page.locator('button[aria-label*="Close" i]').last,
        ]:
            try:
                if locator.count() and locator.last.is_visible(timeout=1000) and locator.last.is_enabled(timeout=1000):
                    locator.last.click(timeout=3000)
                    page.wait_for_timeout(1000)
                    return
            except PlaywrightError:
                pass
        try:
            page.keyboard.press("Escape")
            page.wait_for_timeout(1000)
        except PlaywrightError:
            pass


def find_composer(page: Page):
    for selector in COMPOSER_SELECTORS:
        locator = page.locator(selector).last
        try:
            if locator.count() and locator.is_visible(timeout=2000):
                return locator
        except PlaywrightError:
            pass
    raise RuntimeError("Could not find ChatGPT message composer.")


def composer_text(page: Page) -> str:
    try:
        return page.locator("#prompt-textarea").inner_text(timeout=1000).strip()
    except PlaywrightError:
        return ""


def send_message(page: Page, message: str) -> None:
    dismiss_transient_modal(page)
    composer = find_composer(page)
    for attempt in range(4):
        try:
            composer.click(timeout=5000)
            composer.fill(message, timeout=10000)
            break
        except PlaywrightError:
            dismiss_transient_modal(page)
            if attempt == 3:
                composer.click(timeout=2000, force=True)
                composer.fill(message, timeout=10000)
            page.wait_for_timeout(1000)
    for attempt in range(8):
        dismiss_transient_modal(page)
        for selector in SEND_SELECTORS:
            locator = page.locator(selector).last
            try:
                if locator.count() and locator.is_visible(timeout=1000) and locator.is_enabled(timeout=1000):
                    locator.click(timeout=10000)
                    page.wait_for_timeout(2000)
                    if message.lower() not in composer_text(page).lower():
                        return
            except PlaywrightError:
                pass
        page.keyboard.press("Meta+Enter" if attempt in (2, 5) else "Enter")
        page.wait_for_timeout(3000)
        if message.lower() not in composer_text(page).lower():
            return
    raise RuntimeError("Message did not submit; it is still present in the composer.")


def wait_until_response_done(page: Page, timeout_minutes: int, prompt_number: int) -> None:
    deadline = time.time() + timeout_minutes * 60
    saw_stop = False
    while time.time() < deadline:
        close_non_chatgpt_tabs(page.context, page)
        dismiss_transient_modal(page)
        if phrase_visible(page, CONTEXT_LIMIT_PHRASES):
            raise RuntimeError("ChatGPT appears to be warning about context length.")
        blocking = phrase_visible(page, BLOCKING_PHRASES)
        if blocking:
            raise BlockingNoticeError(f"ChatGPT appears blocked by: {blocking!r}")
        stop_visible = False
        for selector in STOP_SELECTORS:
            try:
                loc = page.locator(selector).last
                if loc.count() and loc.is_visible(timeout=300):
                    stop_visible = True
                    break
            except PlaywrightError:
                pass
        send_enabled = False
        for selector in SEND_SELECTORS:
            try:
                loc = page.locator(selector).last
                if loc.count() and loc.is_visible(timeout=300) and loc.is_enabled(timeout=300):
                    send_enabled = True
                    break
            except PlaywrightError:
                pass
        text = page_text(page)
        final_phrase = f"prompt {prompt_number} completed" in text or f"prompt {prompt_number} could not be completed" in text
        if stop_visible:
            saw_stop = True
        if final_phrase and not stop_visible:
            page.wait_for_timeout(2000)
            return
        if saw_stop and send_enabled:
            page.wait_for_timeout(2000)
            return
        page.wait_for_timeout(2000)
    raise TimeoutError(f"ChatGPT did not finish within {timeout_minutes} minutes.")


def click_download_excel(page: Page, prompt_number: int, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    controls = [
        page.get_by_role("link", name=re.compile(r".*\.xlsx|download.*", re.I)),
        page.get_by_role("button", name=re.compile(r".*\.xlsx|download.*", re.I)),
        page.locator('a[href*=".xlsx"]').last,
        page.locator("[download]").last,
        page.locator('button[aria-label*="Download" i]').last,
        page.locator('a[aria-label*="Download" i]').last,
    ]
    errors: list[str] = []
    for control in controls:
        try:
            if control.count() == 0:
                continue
            item = control.last if hasattr(control, "last") else control
            if not item.is_visible(timeout=2000):
                continue
            with page.expect_download(timeout=8000) as download_info:
                item.click(timeout=10000)
            download = download_info.value
            suggested = download.suggested_filename or f"prompt_{prompt_number:03d}.xlsx"
            if not suggested.lower().endswith(".xlsx"):
                suggested = f"prompt_{prompt_number:03d}_{suggested}.xlsx"
            target = output_dir / f"prompt_{prompt_number:03d}_{int(time.time())}_{suggested}"
            download.save_as(str(target))
            return target
        except Exception as exc:
            errors.append(str(exc))
    raise RuntimeError(f"Could not find/click an Excel download control. Last errors: {errors[-3:]}")


def upload_file(page: Page, file_path: Path) -> None:
    file_path = file_path.expanduser().resolve()
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    inputs = page.locator('input[type="file"]')
    if inputs.count():
        inputs.first.set_input_files(str(file_path))
        page.wait_for_timeout(12000)
        return
    for candidate in [
        page.get_by_role("button", name=re.compile(r"attach|upload|add|file|\+", re.I)),
        page.locator('button[aria-label*="Attach" i]').last,
        page.locator('button[aria-label*="Upload" i]').last,
    ]:
        try:
            if candidate.count() == 0:
                continue
            item = candidate.last if hasattr(candidate, "last") else candidate
            if not item.is_visible(timeout=2000):
                continue
            with page.expect_file_chooser(timeout=10000) as chooser:
                item.click(timeout=10000)
            chooser.value.set_files(str(file_path))
            page.wait_for_timeout(12000)
            return
        except PlaywrightError:
            pass
    raise RuntimeError(f"Could not upload {file_path}.")


def start_new_chat(page: Page) -> None:
    old_url = page.url
    try:
        page.goto("https://chatgpt.com/", wait_until="commit", timeout=15000)
    except PlaywrightTimeoutError:
        print("New-chat navigation did not fully confirm; checking state.")
    page.wait_for_timeout(2000)
    if not is_conversation_page(page):
        print(f"New chat screen opened: {old_url} -> {page.url}")
        return
    raise RuntimeError(f"Could not verify that a new ChatGPT chat opened; still at {page.url}")


def wait_for_cdp(port: int, timeout_seconds: int = 30) -> None:
    deadline = time.time() + timeout_seconds
    while time.time() < deadline:
        try:
            with urllib.request.urlopen(f"http://127.0.0.1:{port}/json/version", timeout=1) as response:
                if response.status == 200:
                    return
        except Exception:
            time.sleep(0.5)
    raise TimeoutError(f"Chrome remote debugging port {port} did not become ready.")


def launch_context(playwright, args: argparse.Namespace):
    chrome_data = Path(args.chrome_data_dir).expanduser().resolve()
    chrome_data.mkdir(parents=True, exist_ok=True)
    chrome_app = Path("/Applications/Google Chrome.app/Contents/MacOS/Google Chrome")
    cmd = [
        str(chrome_app),
        f"--remote-debugging-port={args.remote_debugging_port}",
        f"--user-data-dir={chrome_data}",
        f"--profile-directory={args.chrome_profile}",
        "--disable-blink-features=AutomationControlled",
        "--no-first-run",
        "https://chatgpt.com/",
    ]
    subprocess.Popen(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    wait_for_cdp(args.remote_debugging_port)
    browser = playwright.chromium.connect_over_cdp(f"http://127.0.0.1:{args.remote_debugging_port}")
    if not browser.contexts:
        raise RuntimeError("Connected to Chrome, but no browser context was available.")
    return browser.contexts[0]


def message_for_prompt(prompt_number: int) -> str:
    return f"fill prompt {prompt_number} and audit. Return the full workbook with all original sheets and rows preserved, not a patch/subset workbook."


def missing_workbook_message(prompt_number: int) -> str:
    return f"You completed Prompt {prompt_number}, but I do not see a downloadable Excel workbook attachment. Please return the full updated .xlsx workbook now, with all original sheets and rows preserved."


def wrong_prompt_workbook_message(prompt_number: int) -> str:
    return f"The workbook you just returned did not fill Prompt {prompt_number}; it appears to be for a previous prompt or left the requested row as Not Started. Please correct this now: fill only Prompt {prompt_number}, audit it, preserve all prior rows/sheets/formulas exactly, and return the full updated .xlsx workbook attachment."


def should_rotate_chat(progress: Progress, prompts_per_chat: int) -> bool:
    return progress.next_prompt - progress.current_chat_start_prompt >= prompts_per_chat


def pause_for_user(message: str) -> None:
    print("\n" + message)
    try:
        input("Press Enter to continue... ")
    except EOFError:
        raise SystemExit(2)


def run(args: argparse.Namespace) -> None:
    Path(args.output_dir).mkdir(parents=True, exist_ok=True)
    Path(args.backup_dir).mkdir(parents=True, exist_ok=True)
    progress_path = Path(args.progress)
    progress = load_or_create_progress(args)
    save_progress(progress_path, progress)
    with sync_playwright() as playwright:
        context = launch_context(playwright, args)
        install_tab_guard(context)
        page = chatgpt_page(context)
        close_non_chatgpt_tabs(context, page)
        print("\nChrome is open.")
        if progress.next_prompt == args.start_prompt and not progress.completed_prompts and not args.no_initial_pause:
            print("First run: sign in, choose the model, upload the workbook, then return here.")
            print(f"Workbook: {progress.latest_excel_file}")
            pause_for_user("After the workbook is uploaded and ChatGPT is ready, press Enter.")
        else:
            print(f"Resuming at Prompt {progress.next_prompt}; opening a fresh chat and uploading the latest workbook.")
            start_new_chat(page)
            upload_file(page, Path(progress.latest_excel_file))
            progress.current_chat_start_prompt = progress.next_prompt
            save_progress(progress_path, progress)

        filled = filled_prompt_numbers(Path(args.canonical_excel))
        while progress.next_prompt <= args.end_prompt:
            page = chatgpt_page(context)
            prompt_number = progress.next_prompt
            if prompt_number in filled:
                print(f"Prompt {prompt_number} is already filled in the canonical workbook; skipping.")
                progress.completed_prompts = sorted(set(progress.completed_prompts + [prompt_number]))
                progress.next_prompt += 1
                progress.updated_at = time.time()
                save_progress(progress_path, progress)
                continue
            if should_rotate_chat(progress, args.prompts_per_chat):
                print(f"\nOpening a new chat before Prompt {prompt_number}.")
                start_new_chat(page)
                upload_file(page, Path(progress.latest_excel_file))
                progress.current_chat_start_prompt = prompt_number
                save_progress(progress_path, progress)

            print(f"\nSending: {message_for_prompt(prompt_number)}")
            try:
                send_message(page, message_for_prompt(prompt_number))
                while True:
                    try:
                        wait_until_response_done(page, args.timeout_minutes, prompt_number)
                        break
                    except BlockingNoticeError as exc:
                        print(exc)
                        print(f"Waiting {args.blocking_wait_minutes} minutes before retrying...")
                        time.sleep(args.blocking_wait_minutes * 60)
                try:
                    downloaded = click_download_excel(page, prompt_number, Path(args.output_dir))
                except Exception as exc:
                    print(f"Could not download workbook after first response: {exc}")
                    send_message(page, missing_workbook_message(prompt_number))
                    wait_until_response_done(page, args.timeout_minutes, prompt_number)
                    downloaded = click_download_excel(page, prompt_number, Path(args.output_dir))
            except Exception as exc:
                print(f"\nPrompt {prompt_number} got stuck or failed: {exc}")
                pause_for_user("Fix the ChatGPT page manually if needed, then press Enter to retry this prompt.")
                continue

            for attempt in range(2):
                try:
                    canonical = merge_download_into_canonical(downloaded, Path(args.canonical_excel), prompt_number, Path(args.backup_dir))
                    break
                except RuntimeError as exc:
                    if attempt == 1:
                        raise
                    print(f"Downloaded workbook failed validation for Prompt {prompt_number}: {exc}")
                    print("Asking ChatGPT to correct the workbook for the same prompt, then retrying once.")
                    send_message(page, wrong_prompt_workbook_message(prompt_number))
                    wait_until_response_done(page, args.timeout_minutes, prompt_number)
                    downloaded = click_download_excel(page, prompt_number, Path(args.output_dir))

            validate_workbook_prompt_filled(__import__("openpyxl").load_workbook(canonical, data_only=False, read_only=True), prompt_number, "Canonical workbook")
            filled.add(prompt_number)
            progress.latest_excel_file = str(canonical.resolve())
            progress.completed_prompts = sorted(set(progress.completed_prompts + [prompt_number]))
            progress.next_prompt = prompt_number + 1
            progress.updated_at = time.time()
            save_progress(progress_path, progress)
            print(f"Prompt {prompt_number} complete. Downloaded: {downloaded}")
            print(f"Merged Prompt {prompt_number} into canonical workbook: {canonical}")
            close_non_chatgpt_tabs(context, page)
        print("\nAll requested prompts completed.")
        context.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Automate ChatGPT in real Chrome for sequential Excel prompt filling.")
    parser.add_argument("--initial-excel", default=str(DEFAULT_WORKBOOK))
    parser.add_argument("--canonical-excel", default=str(DEFAULT_WORKBOOK))
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUTS))
    parser.add_argument("--backup-dir", default=str(ROOT / "canonical_backups"))
    parser.add_argument("--progress", default=str(DEFAULT_PROGRESS))
    parser.add_argument("--start-prompt", type=int, default=1)
    parser.add_argument("--end-prompt", type=int, default=670)
    parser.add_argument("--prompts-per-chat", type=int, default=10)
    parser.add_argument("--timeout-minutes", type=int, default=15)
    parser.add_argument("--blocking-wait-minutes", type=int, default=20)
    parser.add_argument("--chrome-data-dir", default=str(ROOT / ".chrome_automation_profile"))
    parser.add_argument("--chrome-profile", default="Default")
    parser.add_argument("--remote-debugging-port", type=int, default=9222)
    parser.add_argument("--no-initial-pause", action="store_true")
    parser.add_argument("--resume", action="store_true")
    return parser


def main() -> int:
    run(build_parser().parse_args())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
